#!/usr/bin/env python3

import argparse
import os
import json
import csv
import re
import string
from datetime import datetime
from abc import ABC, abstractmethod
from pathlib import Path
from typing import Dict, Any, Optional
import hashlib

try:
    import google.generativeai as genai
except ImportError:
    print("Error: google-generativeai not installed. Run: pip install google-generativeai")
    exit(1)

try:
    from dotenv import load_dotenv
    load_dotenv()  # Load .env file if present
except ImportError:
    pass  # dotenv is optional

try:
    import easyocr
except ImportError:
    easyocr = None  # Will be handled gracefully in ImageOCRReader

try:
    import openpyxl
except ImportError:
    openpyxl = None  # Will be handled gracefully in ExcelReader

try:
    import pdfplumber
except ImportError:
    pdfplumber = None  # Will be handled gracefully in PDFReader

try:
    import mammoth
except ImportError:
    mammoth = None  # Will be handled gracefully in WordDocReader


# Utility functions for primary key generation
def generate_file_hash(file_path: str, length: int = 32) -> str:
    """Generate hash of file content for unique identification"""
    try:
        with open(file_path, 'rb') as f:
            return hashlib.md5(f.read()).hexdigest()[:length]
    except Exception as e:
        raise ValueError(f"Cannot generate file hash for {file_path}: {str(e)}")


class DocumentValidator:
    """Configuration-driven data validation for extracted document fields."""
    
    def __init__(self, required_fields: list[str], optional_fields: list[str] = None, field_types: Dict[str, str] = None):
        """
        Initialize validator with field requirements.
        
        Args:
            required_fields: Fields that must be present and non-empty
            optional_fields: Fields that may be present (for documentation)
            field_types: Data type validation for specific fields (e.g., {"Total": "currency", "Date": "date"})
        """
        self.required_fields = required_fields
        self.optional_fields = optional_fields or []
        self.field_types = field_types or {}
    
    def _normalize_currency(self, value: str) -> str:
        """Normalize currency field - remove symbols, commas, whitespace"""
        if not isinstance(value, str):
            return value
        
        # Remove common currency symbols and whitespace
        clean_value = value.replace('$', '').replace('€', '').replace('£', '').replace(',', '').strip()
        return clean_value
    
    def _validate_currency(self, normalized_value: str) -> tuple[bool, str]:
        """Validate normalized currency field - ensure numeric"""
        if not isinstance(normalized_value, str):
            return False, "Currency must be text format"
        
        try:
            float_val = float(normalized_value)
            return True, "Valid currency"
        except ValueError:
            return False, f"Invalid currency format: {normalized_value}"
    
    def _normalize_date(self, value: str) -> str:
        """Normalize date field - basic whitespace cleanup"""
        if not isinstance(value, str):
            return value
        
        return value.strip()
    
    def _validate_date(self, normalized_value: str) -> tuple[bool, str]:
        """Validate normalized date field - check date patterns"""
        if not isinstance(normalized_value, str):
            return False, "Date must be text format"
        
        value = value.strip()
        if not value:
            return False, "Date cannot be empty"
        
        # More comprehensive date patterns
        date_patterns = [
            r'\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b',        # MM/DD/YYYY or MM-DD-YYYY
            r'\b\d{2,4}[/-]\d{1,2}[/-]\d{1,2}\b',        # YYYY/MM/DD or YYYY-MM-DD  
            r'\b[A-Za-z]+ \d{1,2},? \d{4}\b',            # January 15, 2024 or January 15 2024
            r'\b\d{1,2} [A-Za-z]+ \d{4}\b',              # 15 January 2024
            r'\b\d{4}-\d{2}-\d{2}\b',                     # ISO format 2024-01-15
            r'\b\d{2}\.\d{2}\.\d{2,4}\b',                # European format 15.01.2024
        ]
        
        for pattern in date_patterns:
            if re.search(pattern, normalized_value):
                # Additional validation - reject obvious non-dates
                if re.search(r'\d{5,}', normalized_value):  # Too many consecutive digits
                    continue
                return True, "Valid date format"
        
        return False, f"Invalid date format: {normalized_value}"
    
    def _normalize_text(self, value: str) -> str:
        """Normalize text field - strip whitespace"""
        if not isinstance(value, str):
            return value
        
        return value.strip()
    
    def _validate_text(self, normalized_value: str) -> tuple[bool, str]:
        """Validate normalized text field - check emptiness"""
        if not isinstance(normalized_value, str):
            return False, "Text must be string format"
        
        if not normalized_value:
            return False, "Text cannot be empty"
        
        return True, "Valid text"
    
    def _normalize_alphanumeric(self, value: str) -> str:
        """Normalize alphanumeric field - strip whitespace"""
        if not isinstance(value, str):
            return value
        
        return value.strip()
    
    def _validate_alphanumeric(self, normalized_value: str) -> tuple[bool, str]:
        """Validate normalized alphanumeric field - check patterns and length"""
        if not isinstance(normalized_value, str):
            return False, "ID must be string format"
        
        if not normalized_value:
            return False, "ID cannot be empty"
        
        # Allow letters, numbers, hyphens, underscores, spaces
        if not re.match(r'^[A-Za-z0-9\-_\s]+$', normalized_value):
            return False, f"ID contains invalid characters: {normalized_value}"
        
        if len(normalized_value) > 50:  # Reasonable max length for IDs
            return False, "ID exceeds maximum length (50 characters)"
        
        return True, "Valid ID format"
    
    def validate(self, data: Dict[str, Any]) -> tuple[bool, str]:
        """
        Validate extracted data against field requirements.
        
        Args:
            data: Extracted document data
            
        Returns:
            (success, message): Validation result and message
        """
        # Step 1: Check required fields are present
        missing_fields = []
        for field in self.required_fields:
            if field not in data or data[field] is None:
                missing_fields.append(field)
        
        if missing_fields:
            return False, f"Missing required fields: {', '.join(missing_fields)}"
        
        # Step 2: Normalize and validate field types (if specified)
        for field, expected_type in self.field_types.items():
            if field in data and data[field] is not None:
                raw_value = data[field]
                
                # Step 2a: Normalize the field value
                if expected_type == "currency":
                    normalized_value = self._normalize_currency(raw_value)
                elif expected_type == "date":
                    normalized_value = self._normalize_date(raw_value)
                elif expected_type == "text":
                    normalized_value = self._normalize_text(raw_value)
                elif expected_type == "alphanumeric":
                    normalized_value = self._normalize_alphanumeric(raw_value)
                else:
                    continue  # Unknown type, skip normalization and validation
                
                # Step 2b: Validate the normalized value
                if expected_type == "currency":
                    is_valid, msg = self._validate_currency(normalized_value)
                elif expected_type == "date":
                    is_valid, msg = self._validate_date(normalized_value)
                elif expected_type == "text":
                    is_valid, msg = self._validate_text(normalized_value)
                elif expected_type == "alphanumeric":
                    is_valid, msg = self._validate_alphanumeric(normalized_value)
                
                if not is_valid:
                    return False, f"Invalid {expected_type} for {field}: {msg}"
        
        return True, "Validation passed"


class DocumentProcessor(ABC):
    # Extraction is task-focused and intentionally lossy - we only extract
    # the specific fields needed for the task at hand, not all document content
    @abstractmethod
    def get_extracted_fields(self) -> list[str]:
        pass
    
    @abstractmethod
    def extract_fields(self, content: str) -> Dict[str, Any]:
        pass


class DocumentAction(ABC):
    def __init__(self, primary_key_field: Optional[str] = None):
        """
        Initialize action with optional primary key for entity validation.
        
        Args:
            primary_key_field: Field name for business entity validation. None if no field-based validation.
        """
        self.primary_key_field = primary_key_field
    
    def generate_file_key(self, file_path: str) -> str:
        """Generate file hash key for file-level deduplication (always used)."""
        return generate_file_hash(file_path)
    
    def generate_primary_key(self, data: Dict[str, Any]) -> Optional[str]:
        """Generate business entity key for data quality validation (when available)."""
        if self.primary_key_field and self.primary_key_field in data:
            key_value = data[self.primary_key_field]
            if key_value is None:
                return None
            
            # Convert to string and clean
            key_str = str(key_value).strip()
            if not key_str:
                return None
            
            # Sanitize for filename safety (remove/replace problematic characters)
            safe_chars = string.ascii_letters + string.digits + '-_.'
            sanitized = ''.join(c if c in safe_chars else '_' for c in key_str)
            
            # Limit length to prevent filesystem issues
            max_length = 100
            if len(sanitized) > max_length:
                sanitized = sanitized[:max_length].rstrip('_')
            
            return sanitized if sanitized else None
        return None
    
    @abstractmethod
    def generate_filename(self, doc_type: str, data: Dict[str, Any], file_path: str, key: str) -> str:
        """Generate output filename. Override for different naming strategies."""
        pass
    
    def check_file_duplicates(self, file_key: str, output_path: str) -> bool:
        """Check if this file was already processed. Override for custom logic."""
        return False  # Default: no file-level deduplication
    
    def check_primary_key_duplicates(self, entity_key: Optional[str], output_path: str) -> bool:
        """Check if this business entity already exists. Override for custom logic."""
        return False  # Default: no primary key-level deduplication
    
    def check_duplicates(self, data: Dict[str, Any], file_path: str, key: str, output_path: str) -> bool:
        """Check if this should be skipped as duplicate. Override for custom logic."""
        return False  # Default: no deduplication (backward compatibility)
    
    @abstractmethod
    def validate_data(self, data: Dict[str, Any]) -> tuple[bool, str]:
        """Validate data meets action requirements. Returns (success, error_msg)"""
        pass
    
    @abstractmethod 
    def execute(self, data: Dict[str, Any], file_path: str) -> tuple[bool, str]:
        """Execute action. Returns (success, result_msg)"""
        pass


class APInvoiceCSVAction(DocumentAction):
    def __init__(self):
        super().__init__(primary_key_field="Invoice #")
        self.validator = DocumentValidator(
            required_fields=["Vendor", "Invoice #", "Date Due", "Total"],
            field_types={
                "Vendor": "text",
                "Invoice #": "alphanumeric",
                "Date Due": "date",
                "Total": "currency"
            }
        )
    
    def generate_filename(self, doc_type: str, data: Dict[str, Any], file_path: str, key: str) -> str:
        """Generate compiled CSV file (compile pattern)"""
        return f"{doc_type}_log.csv"
    
    def check_file_duplicates(self, file_key: str, output_path: str) -> bool:
        """Check if file hash already exists in CSV (primary deduplication)"""
        return self._is_file_duplicate(file_key, output_path)
    
    def check_primary_key_duplicates(self, entity_key: Optional[str], output_path: str) -> bool:
        """Check if business key already exists in CSV (data quality validation)"""
        if entity_key is None:
            return False
        return self._is_entity_duplicate(entity_key, output_path)
    
    def check_duplicates(self, data: Dict[str, Any], file_path: str, key: str, output_path: str) -> bool:
        """Legacy method for backward compatibility"""
        file_key = self.generate_file_key(file_path)
        return self.check_file_duplicates(file_key, output_path)
    
    def validate_data(self, data: Dict[str, Any]) -> tuple[bool, str]:
        """Validate AP Invoice data has all required fields"""
        return self.validator.validate(data)
    
    def _is_file_duplicate(self, file_key: str, csv_path: str) -> bool:
        """Check if file hash already exists in CSV file"""
        if not os.path.exists(csv_path):
            return False  # No file means no duplicates
        
        try:
            with open(csv_path, 'r', encoding='utf-8') as csvfile:
                reader = csv.reader(csvfile)
                header = next(reader, None)  # Skip header row
                if not header:
                    return False  # Empty file
                
                # Find the File Hash column index
                try:
                    hash_index = header.index("File Hash")
                except ValueError:
                    return False  # File Hash column doesn't exist
                
                # Check each row for the file hash
                for row in reader:
                    if len(row) > hash_index and row[hash_index] == file_key:
                        return True  # Found duplicate
                        
            return False  # No duplicate found
            
        except Exception:
            return False  # Error reading file, assume no duplicate
    
    def _is_entity_duplicate(self, entity_key: str, csv_path: str) -> bool:
        """Check if business entity already exists in CSV file"""
        if not os.path.exists(csv_path):
            return False  # No file means no duplicates
        
        try:
            with open(csv_path, 'r', encoding='utf-8') as csvfile:
                reader = csv.reader(csvfile)
                header = next(reader, None)  # Skip header row
                if not header:
                    return False  # Empty file
                
                # Find the column index for our primary key
                try:
                    key_index = header.index(self.primary_key_field)
                except ValueError:
                    return False  # Primary key column doesn't exist
                
                # Check each row for the entity key
                for row in reader:
                    if len(row) > key_index and row[key_index] == entity_key:
                        return True  # Found duplicate
                        
            return False  # No duplicate found
            
        except Exception:
            return False  # Error reading file, assume no duplicate
    
    def _is_duplicate(self, data: Dict[str, Any], file_path: str, csv_path: str) -> bool:
        """Legacy method: Check if primary key already exists in CSV file"""
        entity_key = self.generate_primary_key(data)
        if not entity_key:
            return False
        return self._is_entity_duplicate(entity_key, csv_path)
    
    def execute(self, data: Dict[str, Any], file_path: str) -> tuple[bool, str]:
        """Append AP Invoice data to CSV using dual-key system"""
        file_key = self.generate_file_key(file_path)
        entity_key = self.generate_primary_key(data)
        csv_path = self.generate_filename("ap_invoice", data, file_path, file_key)
        
        # Primary check: File-level deduplication (robust)
        if self.check_file_duplicates(file_key, csv_path):
            return True, "duplicate"
        
        # Secondary check: Primary key-level validation (data quality)
        if entity_key and self.check_primary_key_duplicates(entity_key, csv_path):
            # Log entity duplication as a warning but don't block processing
            print(f"WARNING: Business entity {entity_key} already exists but from different file - {file_path}")
        
        try:
            # Check if file exists to determine if we need headers
            file_exists = os.path.exists(csv_path)
            
            # Define field order for CSV (include File Hash)
            field_order = ["Vendor", "Invoice #", "Date Due", "Total", "File Hash"]
            
            with open(csv_path, 'a', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write headers if new file
                if not file_exists:
                    writer.writerow(field_order)
                
                # Write data row in correct order (include file hash)
                row = [data["Vendor"], data["Invoice #"], data["Date Due"], data["Total"], file_key]
                writer.writerow(row)
            
            return True, "appended"
            
        except Exception as e:
            return False, f"Failed to write to CSV: {str(e)}"


class PurchaseOrderJSONAction(DocumentAction):
    def __init__(self):
        super().__init__(primary_key_field="PO Number")
        self.validator = DocumentValidator(
            required_fields=["PO Number", "Vendor", "Order Date", "Delivery Date", "Total Amount"],
            field_types={
                "PO Number": "alphanumeric",
                "Vendor": "text",
                "Order Date": "date",
                "Delivery Date": "date",
                "Total Amount": "currency"
            }
        )
    
    def generate_filename(self, doc_type: str, data: Dict[str, Any], file_path: str, key: str) -> str:
        """Generate individual JSON file using file key (transform pattern)"""
        return f"{doc_type}_{key}.json"
    
    def validate_data(self, data: Dict[str, Any]) -> tuple[bool, str]:
        """Validate Purchase Order data has all required fields"""
        return self.validator.validate(data)
    
    def execute(self, data: Dict[str, Any], file_path: str) -> tuple[bool, str]:
        """Export Purchase Order data to individual JSON file using dual-key system"""
        file_key = self.generate_file_key(file_path)
        entity_key = self.generate_primary_key(data)
        filename = self.generate_filename("purchase_order", data, file_path, file_key)
        
        # Check for duplicates (returns False by default for JSON files)
        if self.check_duplicates(data, file_path, file_key, filename):
            return True, "duplicate"
        
        # Generate structured JSON
        json_output = {
            "file_hash": file_key,
            "po_number": entity_key,  # Keep business key for data integrity
            "vendor": data["Vendor"],
            "dates": {
                "order_date": data["Order Date"],
                "delivery_date": data["Delivery Date"]
            },
            "financial": {
                "total_amount": data["Total Amount"]
            },
            "source_file": os.path.basename(file_path),
            "processed_timestamp": datetime.now().isoformat()
        }
        
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(json_output, f, indent=2)
            return True, "json_created"
        except Exception as e:
            return False, f"Failed to write JSON ({filename}): {str(e)}"


class ReceiptJSONAction(DocumentAction):
    def __init__(self):
        super().__init__(primary_key_field=None)  # No business entity validation for receipts
        self.validator = DocumentValidator(
            required_fields=["Merchant", "Date", "Amount", "Payment Method"],
            optional_fields=["Category"],
            field_types={
                "Merchant": "text",
                "Date": "date", 
                "Amount": "currency",
                "Payment Method": "text",
                "Category": "text"
            }
        )
    
    def generate_filename(self, doc_type: str, data: Dict[str, Any], file_path: str, key: str) -> str:
        """Generate individual JSON file using hash (transform pattern)"""
        return f"{doc_type}_{key}.json"
    
    def validate_data(self, data: Dict[str, Any]) -> tuple[bool, str]:
        """Validate Receipt data has all required fields"""
        return self.validator.validate(data)
    
    def execute(self, data: Dict[str, Any], file_path: str) -> tuple[bool, str]:
        """Export Receipt data to individual JSON file using dual-key system"""
        file_key = self.generate_file_key(file_path)
        filename = self.generate_filename("receipt", data, file_path, file_key)
        
        # Check for duplicates (returns False by default for JSON files)
        if self.check_duplicates(data, file_path, file_key, filename):
            return True, "duplicate"
        
        # Generate structured JSON
        json_output = {
            "merchant": data["Merchant"],
            "date": data["Date"],
            "amount": data["Amount"],
            "payment_method": data["Payment Method"],
            "source_file": os.path.basename(file_path),
            "file_hash": file_key,
            "processed_timestamp": datetime.now().isoformat()
        }
        
        # Only include category if present in data (don't infer)
        if "Category" in data and data["Category"]:
            json_output["category"] = data["Category"]
        
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(json_output, f, indent=2)
            return True, "json_created"
        except Exception as e:
            return False, f"Failed to write JSON ({filename}): {str(e)}"


class BankStatementCSVAction(DocumentAction):
    def __init__(self):
        super().__init__(primary_key_field=None)  # No business entity validation for bank statements
        self.validator = DocumentValidator(
            required_fields=["Amount", "Date", "Description"],
            field_types={
                "Amount": "currency",
                "Date": "date",
                "Description": "text"
            }
        )
    
    def generate_filename(self, doc_type: str, data: Dict[str, Any], file_path: str, key: str) -> str:
        """Generate compiled CSV file (compile pattern)"""
        return f"{doc_type}_transactions.csv"
    
    def check_file_duplicates(self, file_key: str, output_path: str) -> bool:
        """Check if file hash already exists in CSV (primary deduplication)"""
        return self._is_file_duplicate(file_key, output_path)
    
    def check_duplicates(self, data: Dict[str, Any], file_path: str, key: str, output_path: str) -> bool:
        """Legacy method for backward compatibility"""
        file_key = self.generate_file_key(file_path)
        return self.check_file_duplicates(file_key, output_path)
    
    def validate_data(self, data: Dict[str, Any]) -> tuple[bool, str]:
        """Validate Bank Statement data has all required fields"""
        return self.validator.validate(data)
    
    def _is_file_duplicate(self, file_key: str, csv_path: str) -> bool:
        """Check if file hash already exists in CSV file"""
        if not os.path.exists(csv_path):
            return False  # No file means no duplicates
        
        try:
            with open(csv_path, 'r', encoding='utf-8') as csvfile:
                reader = csv.reader(csvfile)
                header = next(reader, None)  # Skip header row
                if not header:
                    return False  # Empty file
                
                # Find the File Hash column index
                try:
                    hash_index = header.index("File Hash")
                except ValueError:
                    return False  # File Hash column doesn't exist
                
                # Check each row for the file hash
                for row in reader:
                    if len(row) > hash_index and row[hash_index] == file_key:
                        return True  # Found duplicate
                        
            return False  # No duplicate found
            
        except Exception:
            return False  # Error reading file, assume no duplicate
    
    def _is_duplicate(self, data: Dict[str, Any], file_path: str, csv_path: str) -> bool:
        """Legacy method for backward compatibility"""
        file_key = self.generate_file_key(file_path)
        return self._is_file_duplicate(file_key, csv_path)
    
    def execute(self, data: Dict[str, Any], file_path: str) -> tuple[bool, str]:
        """Append Bank Statement transaction to CSV using dual-key system"""
        file_key = self.generate_file_key(file_path)
        csv_path = self.generate_filename("bank_statement", data, file_path, file_key)
        
        # Check for file duplicates (primary deduplication)
        if self.check_file_duplicates(file_key, csv_path):
            return True, "duplicate"
        
        try:
            # Check if file exists to determine if we need headers
            file_exists = os.path.exists(csv_path)
            
            # Define field order for CSV (include File Hash)
            field_order = ["Date", "Amount", "Description", "File Hash"]
            
            with open(csv_path, 'a', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # Write headers if new file
                if not file_exists:
                    writer.writerow(field_order)
                
                # Write data row in correct order (include file hash)
                row = [data["Date"], data["Amount"], data["Description"], file_key]
                writer.writerow(row)
            
            return True, "appended"
            
        except Exception as e:
            return False, f"Failed to write to CSV: {str(e)}"


def validate_json_response(response_text: str, expected_fields: list[str]) -> tuple[bool, Dict[str, Any], str]:
    """Validate JSON response and check for expected fields."""
    if not response_text or not isinstance(response_text, str):
        return False, {}, "Empty or invalid response"
    
    # Clean markdown formatting from response
    cleaned_text = response_text.strip()
    if cleaned_text.startswith('```json'):
        cleaned_text = cleaned_text[7:]  # Remove ```json
    if cleaned_text.startswith('```'):
        cleaned_text = cleaned_text[3:]   # Remove ```
    if cleaned_text.endswith('```'):
        cleaned_text = cleaned_text[:-3]  # Remove ```
    cleaned_text = cleaned_text.strip()
    
    if not cleaned_text:
        return False, {}, "Response is empty after cleaning"
    
    try:
        data = json.loads(cleaned_text)
        if not isinstance(data, dict):
            return False, {}, "Response is not a JSON object"
        
        # Check all expected fields are present
        missing_fields = []
        empty_fields = []
        
        for field in expected_fields:
            if field not in data:
                missing_fields.append(field)
            elif data[field] is None or (isinstance(data[field], str) and not data[field].strip()):
                empty_fields.append(field)
        
        if missing_fields:
            return False, {}, f"Missing required fields: {', '.join(missing_fields)}"
        
        if empty_fields:
            return False, {}, f"Empty required fields: {', '.join(empty_fields)}"
        
        return True, data, "Success"
    except json.JSONDecodeError as e:
        return False, {}, f"Invalid JSON format: {str(e)}"
    except UnicodeDecodeError as e:
        return False, {}, f"Unicode decode error: {str(e)}"


class GeminiClient:
    def __init__(self):
        api_key = os.getenv('GEMINI_API_KEY')
        if not api_key:
            raise ValueError("GEMINI_API_KEY environment variable not set")
        
        genai.configure(api_key=api_key)
        self.model = genai.GenerativeModel('gemini-2.0-flash-exp')
    
    def extract_fields(self, content: str, expected_fields: list[str]) -> Dict[str, Any]:
        errors = []
        
        for attempt in range(3):
            if attempt == 0:
                prompt = f"""Extract the following fields from this document:
{', '.join(expected_fields)}

Please return the response in valid JSON format.

Document content:
{content}

Return only a JSON object with the extracted values. Use null for missing fields. 
IMPORTANT: Return valid JSON format only."""
            else:
                prompt = f"""You must extract EXACTLY these fields from this document:
{', '.join(expected_fields)}

CRITICAL: Your response must be ONLY valid JSON. No explanations, no markdown formatting, no additional text.

Document content:
{content}

Return a JSON object with these exact field names: {', '.join(expected_fields)}
Use null for any missing values.
RESPOND WITH VALID JSON ONLY."""
            
            try:
                response = self.model.generate_content(prompt)
                is_valid, data, error_msg = validate_json_response(response.text, expected_fields)
                
                if is_valid:
                    return data
                else:
                    errors.append(f"Attempt {attempt + 1}: {error_msg}")
                    
            except Exception as e:
                errors.append(f"Attempt {attempt + 1}: API error - {str(e)}")
        
        # Print detailed error information
        print("ERROR: All 3 extraction attempts failed:")
        for error in errors:
            print(f"  - {error}")
        return None


class APInvoiceProcessor(DocumentProcessor):
    def __init__(self, gemini_client: GeminiClient):
        self.gemini_client = gemini_client
        self.action = APInvoiceCSVAction()
    
    def get_extracted_fields(self) -> list[str]:
        return ["Vendor", "Invoice #", "Date Due", "Total"]
    
    def extract_fields(self, content: str) -> Dict[str, Any]:
        return self.gemini_client.extract_fields(content, self.get_extracted_fields())
    
    def execute_action(self, data: Dict[str, Any], file_path: str) -> tuple[bool, str]:
        """Execute the post-extraction action"""
        # First validate the data
        is_valid, msg = self.action.validate_data(data)
        if not is_valid:
            return False, f"Validation failed: {msg}"
        
        # Execute the action
        return self.action.execute(data, file_path)


class PurchaseOrderProcessor(DocumentProcessor):
    def __init__(self, gemini_client: GeminiClient):
        self.gemini_client = gemini_client
        self.action = PurchaseOrderJSONAction()
    
    def get_extracted_fields(self) -> list[str]:
        return ["PO Number", "Vendor", "Order Date", "Delivery Date", "Total Amount"]
    
    def extract_fields(self, content: str) -> Dict[str, Any]:
        return self.gemini_client.extract_fields(content, self.get_extracted_fields())
    
    def execute_action(self, data: Dict[str, Any], file_path: str) -> tuple[bool, str]:
        """Execute the post-extraction action"""
        # First validate the data
        is_valid, msg = self.action.validate_data(data)
        if not is_valid:
            return False, f"Validation failed: {msg}"
        
        # Execute the action
        return self.action.execute(data, file_path)


class ReceiptProcessor(DocumentProcessor):
    def __init__(self, gemini_client: GeminiClient):
        self.gemini_client = gemini_client
        self.action = ReceiptJSONAction()
    
    def get_extracted_fields(self) -> list[str]:
        return ["Merchant", "Date", "Amount", "Payment Method", "Category"]
    
    def extract_fields(self, content: str) -> Dict[str, Any]:
        return self.gemini_client.extract_fields(content, self.get_extracted_fields())
    
    def execute_action(self, data: Dict[str, Any], file_path: str) -> tuple[bool, str]:
        """Execute the post-extraction action"""
        # First validate the data
        is_valid, msg = self.action.validate_data(data)
        if not is_valid:
            return False, f"Validation failed: {msg}"
        
        # Execute the action
        return self.action.execute(data, file_path)


class BankStatementProcessor(DocumentProcessor):
    def __init__(self, gemini_client: GeminiClient):
        self.gemini_client = gemini_client
        self.action = BankStatementCSVAction()
    
    def get_extracted_fields(self) -> list[str]:
        return ["Amount", "Date", "Description"]
    
    def extract_fields(self, content: str) -> Dict[str, Any]:
        return self.gemini_client.extract_fields(content, self.get_extracted_fields())
    
    def execute_action(self, data: Dict[str, Any], file_path: str) -> tuple[bool, str]:
        """Execute the post-extraction action"""
        # First validate the data
        is_valid, msg = self.action.validate_data(data)
        if not is_valid:
            return False, f"Validation failed: {msg}"
        
        # Execute the action
        return self.action.execute(data, file_path)


class ProcessorRegistry:
    # TODO: Add more document types (contracts, etc.)
    _processors = {
        'ap_invoice': APInvoiceProcessor,
        'purchase_order': PurchaseOrderProcessor,
        'receipt': ReceiptProcessor,
        'bank_statement': BankStatementProcessor
    }
    
    @classmethod
    def get_processor(cls, doc_type: str, gemini_client: GeminiClient) -> Optional[DocumentProcessor]:
        processor_class = cls._processors.get(doc_type)
        if processor_class:
            try:
                return processor_class(gemini_client)
            except ValueError as e:
                print(f"Error initializing processor: {e}")
                return None
        return None
    
    @classmethod
    def get_available_types(cls) -> list[str]:
        return list(cls._processors.keys())


class FileReader(ABC):
    """Abstract base class for reading different file types."""
    
    @abstractmethod
    def can_read(self, file_path: str) -> bool:
        """Check if this reader can handle the given file type."""
        pass
    
    @abstractmethod
    def read(self, file_path: str) -> str:
        """Read and return the text content of the file."""
        pass
    
    @abstractmethod
    def get_supported_extensions(self) -> list[str]:
        """Return list of supported file extensions."""
        pass


class TextFileReader(FileReader):
    """Reader for plain text files."""
    
    def can_read(self, file_path: str) -> bool:
        return Path(file_path).suffix.lower() in self.get_supported_extensions()
    
    def read(self, file_path: str) -> str:
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()
    
    def get_supported_extensions(self) -> list[str]:
        return ['.txt', '.md', '.markdown']


class ImageOCRReader(FileReader):
    """Reader for image files using EasyOCR for text extraction."""
    
    def __init__(self):
        self.reader = None
        self._check_dependencies()
    
    def _check_dependencies(self):
        """Check if EasyOCR is available and initialize it lazily."""
        if easyocr is None:
            raise ImportError("EasyOCR not installed. Run: pip install easyocr")
    
    def _get_ocr_reader(self):
        """Lazy initialization of EasyOCR reader."""
        if self.reader is None:
            # Initialize with English language support
            # EasyOCR will download models on first use (~47MB for English)
            self.reader = easyocr.Reader(['en'])
        return self.reader
    
    def can_read(self, file_path: str) -> bool:
        if easyocr is None:
            return False
        return Path(file_path).suffix.lower() in self.get_supported_extensions()
    
    def read(self, file_path: str) -> str:
        try:
            ocr_reader = self._get_ocr_reader()
            
            # Extract text from image
            results = ocr_reader.readtext(file_path)
            
            # Combine all detected text into a single string
            # Each result is [bbox, text, confidence]
            extracted_lines = []
            for bbox, text, confidence in results:
                # Only include text with reasonable confidence (>0.5)
                if confidence > 0.5:
                    extracted_lines.append(text)
            
            if not extracted_lines:
                raise ValueError("No text detected in image or confidence too low")
            
            # Join lines with newlines to preserve document structure
            return '\n'.join(extracted_lines)
            
        except Exception as e:
            raise ValueError(f"Failed to extract text from image {file_path}: {str(e)}")
    
    def get_supported_extensions(self) -> list[str]:
        return ['.png', '.jpg', '.jpeg', '.tiff', '.tif', '.bmp', '.webp']


class CSVReader(FileReader):
    """Reader for CSV files, converts to readable text format."""
    
    def can_read(self, file_path: str) -> bool:
        return Path(file_path).suffix.lower() in self.get_supported_extensions()
    
    def read(self, file_path: str) -> str:
        try:
            # Read CSV and convert to readable text format
            rows = []
            with open(file_path, 'r', encoding='utf-8') as csvfile:
                # Try to detect delimiter automatically
                sample = csvfile.read(1024)
                csvfile.seek(0)
                sniffer = csv.Sniffer()
                
                try:
                    delimiter = sniffer.sniff(sample).delimiter
                except:
                    delimiter = ','  # Default fallback
                
                reader = csv.reader(csvfile, delimiter=delimiter)
                rows = list(reader)
            
            if not rows:
                raise ValueError("CSV file is empty")
            
            # Convert to readable text format for LLM processing
            text_lines = []
            
            # Add header if available
            if rows:
                header = rows[0]
                text_lines.append("CSV Data with columns: " + ", ".join(header))
                text_lines.append("")  # Empty line for separation
                
                # Add data rows with structure
                for i, row in enumerate(rows[1:], 1):
                    if len(row) == len(header):
                        row_text = f"Row {i}:"
                        for col_name, value in zip(header, row):
                            row_text += f" {col_name}: {value},"
                        text_lines.append(row_text.rstrip(','))
                    else:
                        # Handle malformed rows
                        text_lines.append(f"Row {i}: {', '.join(row)}")
                
                # Limit to reasonable number of rows for LLM processing
                if len(text_lines) > 102:  # Header + blank + 100 rows
                    text_lines = text_lines[:102]
                    text_lines.append(f"... ({len(rows)-1-100} more rows truncated)")
            
            return '\n'.join(text_lines)
            
        except Exception as e:
            raise ValueError(f"Failed to read CSV file {file_path}: {str(e)}")
    
    def get_supported_extensions(self) -> list[str]:
        return ['.csv']


class ExcelReader(FileReader):
    """Reader for Excel files using openpyxl, converts to readable text format."""
    
    def __init__(self):
        self._check_dependencies()
    
    def _check_dependencies(self):
        """Check if openpyxl is available."""
        if openpyxl is None:
            raise ImportError("openpyxl not installed. Run: pip install openpyxl")
    
    def can_read(self, file_path: str) -> bool:
        if openpyxl is None:
            return False
        return Path(file_path).suffix.lower() in self.get_supported_extensions()
    
    def read(self, file_path: str) -> str:
        try:
            # Load the workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            text_lines = []
            
            # Process all sheets
            for sheet_idx, sheet_name in enumerate(workbook.sheetnames):
                sheet = workbook[sheet_name]
                
                # Add sheet header if multiple sheets
                if len(workbook.sheetnames) > 1:
                    text_lines.append(f"Sheet: {sheet_name}")
                    text_lines.append("")
                
                # Get all rows with data
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    # Skip completely empty rows
                    if any(cell is not None for cell in row):
                        # Convert None to empty string, everything else to string
                        cleaned_row = [str(cell) if cell is not None else "" for cell in row]
                        rows.append(cleaned_row)
                
                if not rows:
                    text_lines.append("(Empty sheet)")
                    text_lines.append("")
                    continue
                
                # Assume first row is headers
                if rows:
                    headers = rows[0]
                    text_lines.append(f"Excel Data with columns: {', '.join(headers)}")
                    text_lines.append("")
                    
                    # Add data rows with structure (limit to 100 rows per sheet)
                    data_rows = rows[1:101] if len(rows) > 1 else []
                    
                    for i, row in enumerate(data_rows, 1):
                        if len(row) == len(headers):
                            row_text = f"Row {i}:"
                            for col_name, value in zip(headers, row):
                                if value:  # Skip empty values
                                    row_text += f" {col_name}: {value},"
                            text_lines.append(row_text.rstrip(','))
                        else:
                            # Handle mismatched row lengths
                            non_empty_values = [val for val in row if val]
                            if non_empty_values:
                                text_lines.append(f"Row {i}: {', '.join(non_empty_values)}")
                    
                    # Add truncation note if there are more rows
                    if len(rows) > 101:
                        text_lines.append(f"... ({len(rows)-101} more rows truncated)")
                
                text_lines.append("")  # Blank line between sheets
            
            return '\n'.join(text_lines).strip()
            
        except Exception as e:
            raise ValueError(f"Failed to read Excel file {file_path}: {str(e)}")
    
    def get_supported_extensions(self) -> list[str]:
        return ['.xlsx', '.xlsm', '.xltx', '.xltm']


class PDFReader(FileReader):
    """Reader for PDF files using pdfplumber for text extraction."""
    
    def __init__(self):
        self._check_dependencies()
    
    def _check_dependencies(self):
        """Check if pdfplumber is available."""
        if pdfplumber is None:
            raise ImportError("pdfplumber not installed. Run: pip install pdfplumber")
    
    def can_read(self, file_path: str) -> bool:
        if pdfplumber is None:
            return False
        return Path(file_path).suffix.lower() in self.get_supported_extensions()
    
    def read(self, file_path: str) -> str:
        try:
            text_lines = []
            
            with pdfplumber.open(file_path) as pdf:
                total_pages = len(pdf.pages)
                
                if total_pages == 0:
                    raise ValueError("PDF file contains no pages")
                
                # Add document header
                text_lines.append(f"PDF Document ({total_pages} pages)")
                text_lines.append("")
                
                # Process pages (limit to reasonable number for LLM context)
                max_pages = 20  # Reasonable limit for LLM processing
                pages_to_process = min(total_pages, max_pages)
                
                for page_num in range(pages_to_process):
                    page = pdf.pages[page_num]
                    
                    # Extract text from page
                    page_text = page.extract_text()
                    
                    if page_text:
                        # Add page header if multiple pages
                        if total_pages > 1:
                            text_lines.append(f"Page {page_num + 1}:")
                            text_lines.append("")
                        
                        # Clean up the text
                        cleaned_text = page_text.strip()
                        if cleaned_text:
                            text_lines.append(cleaned_text)
                            text_lines.append("")  # Blank line between pages
                    else:
                        # Handle pages with no extractable text (might be images/scans)
                        if total_pages > 1:
                            text_lines.append(f"Page {page_num + 1}: (No extractable text - may contain images)")
                            text_lines.append("")
                
                # Add truncation note if there are more pages
                if total_pages > max_pages:
                    text_lines.append(f"... ({total_pages - max_pages} more pages truncated)")
                
                result = '\n'.join(text_lines).strip()
                
                if not result or result == f"PDF Document ({total_pages} pages)":
                    raise ValueError("No text could be extracted from PDF - document may be image-based or corrupted")
                
                return result
                
        except Exception as e:
            raise ValueError(f"Failed to read PDF file {file_path}: {str(e)}")
    
    def get_supported_extensions(self) -> list[str]:
        return ['.pdf']


class WordDocReader(FileReader):
    """Reader for Word documents using mammoth for text extraction."""
    
    def __init__(self):
        self._check_dependencies()
    
    def _check_dependencies(self):
        """Check if mammoth is available."""
        if mammoth is None:
            raise ImportError("mammoth not installed. Run: pip install mammoth")
    
    def can_read(self, file_path: str) -> bool:
        if mammoth is None:
            return False
        return Path(file_path).suffix.lower() in self.get_supported_extensions()
    
    def read(self, file_path: str) -> str:
        try:
            with open(file_path, "rb") as docx_file:
                result = mammoth.extract_raw_text(docx_file)
                
                # Get the extracted text
                text = result.value
                
                # Check for any warnings or issues
                if result.messages:
                    # Mammoth sometimes reports warnings about unrecognized styles
                    # We'll note serious errors but continue with text extraction
                    error_messages = [msg for msg in result.messages if msg.type == 'error']
                    if error_messages:
                        print(f"Warning: Some issues encountered while reading {file_path}")
                        for msg in error_messages:
                            print(f"  - {msg.message}")
                
                if not text or not text.strip():
                    raise ValueError("No text could be extracted from Word document")
                
                # Clean up the text
                cleaned_text = text.strip()
                
                # Add document header
                lines = [f"Word Document Content:", ""]
                lines.extend(cleaned_text.split('\n'))
                
                return '\n'.join(lines)
                
        except Exception as e:
            if isinstance(e, ValueError) and "No text could be extracted" in str(e):
                raise e
            raise ValueError(f"Failed to read Word document {file_path}: {str(e)}")
    
    def get_supported_extensions(self) -> list[str]:
        return ['.docx']


class FileReaderRegistry:
    """Registry for managing file readers and automatic file type detection."""
    
    def __init__(self):
        self.readers: list[FileReader] = []
        self._register_default_readers()
    
    def _register_default_readers(self):
        """Register built-in file readers."""
        self.readers.append(TextFileReader())
        self.readers.append(CSVReader())
        
        # Register ExcelReader if openpyxl is available
        try:
            self.readers.append(ExcelReader())
        except ImportError:
            pass  # Skip if openpyxl not available
        
        # Register PDFReader if pdfplumber is available
        try:
            self.readers.append(PDFReader())
        except ImportError:
            pass  # Skip if pdfplumber not available
        
        # Register WordDocReader if mammoth is available
        try:
            self.readers.append(WordDocReader())
        except ImportError:
            pass  # Skip if mammoth not available
        
        # Register ImageOCRReader if EasyOCR is available
        try:
            self.readers.append(ImageOCRReader())
        except ImportError:
            pass  # Skip if EasyOCR not available
    
    def register_reader(self, reader: FileReader):
        """Register a new file reader."""
        self.readers.append(reader)
    
    def get_reader(self, file_path: str) -> Optional[FileReader]:
        """Find the first reader that can handle this file type."""
        for reader in self.readers:
            if reader.can_read(file_path):
                return reader
        return None
    
    def get_supported_extensions(self) -> list[str]:
        """Get all supported file extensions across all readers."""
        extensions = []
        for reader in self.readers:
            extensions.extend(reader.get_supported_extensions())
        return sorted(list(set(extensions)))


# Global file reader registry instance
_file_reader_registry = FileReaderRegistry()


def read_file(file_path: str) -> str:
    """Read file content using the appropriate file reader."""
    reader = _file_reader_registry.get_reader(file_path)
    if reader is None:
        file_ext = Path(file_path).suffix.lower()
        supported = _file_reader_registry.get_supported_extensions()
        raise ValueError(f"Unsupported file type: {file_ext}. Supported: {', '.join(supported)}")
    
    return reader.read(file_path)


def main():
    parser = argparse.ArgumentParser(description='Ingest documents')
    parser.add_argument('files', nargs='+', help='Files to ingest')
    # TODO: Make --type optional with auto-detection of document type by default
    parser.add_argument('--type', dest='doc_type', required=True,
                        choices=ProcessorRegistry.get_available_types(),
                        help='Document type')
    
    args = parser.parse_args()
    
    # Initialize shared Gemini client once for all processors
    try:
        gemini_client = GeminiClient()
    except ValueError as e:
        print(f"ERROR: Failed to initialize Gemini client: {e}")
        return
    
    processed = 0
    duplicates = 0
    failed = 0
    
    for file_path in args.files:
        try:
            # Try to read the file first
            try:
                content = read_file(file_path)
            except FileNotFoundError:
                print(f"ERROR: File not found - {file_path}")
                failed += 1
                continue
            except ValueError as e:
                print(f"ERROR: Unsupported file type - {file_path} ({e})")
                failed += 1
                continue
            except Exception as e:
                print(f"ERROR: Could not read file - {file_path} ({e})")
                failed += 1
                continue
            
            # Try to get processor
            processor = ProcessorRegistry.get_processor(args.doc_type, gemini_client)
            if not processor:
                print(f"ERROR: Could not initialize processor for document type '{args.doc_type}' - {file_path}")
                failed += 1
                continue
                
            print(f"Processing {file_path} as {args.doc_type}")
            
            # Try extraction
            extracted_data = processor.extract_fields(content)
            
            if extracted_data is None:
                print(f"ERROR: Field extraction failed - {file_path}")
                failed += 1
            else:
                # Execute post-extraction action (single action per document type)
                # NOTE: Current design implements one action per processor for assignment requirements.
                # Architecture supports extending to action pipelines for production (e.g., csv + webform + email chain)
                # by modifying processors to hold action lists instead of single action instances.
                # TODO: Add --action CLI argument to support action selection (e.g., --action csv, --action json)
                if hasattr(processor, 'execute_action'):
                    action_success, action_msg = processor.execute_action(extracted_data, file_path)
                    if action_success:
                        if action_msg == "duplicate":
                            duplicates += 1
                        else:  # "appended" or other success
                            processed += 1
                    else:
                        print(f"ERROR: Action failed - {file_path}: {action_msg}")
                        failed += 1
                else:
                    # No action defined for this processor type
                    processed += 1
                
        except Exception as e:
            print(f"ERROR: Unexpected error processing {file_path}: {e}")
            failed += 1
    
    total_files = processed + duplicates + failed
    if total_files == 1:
        filename = args.files[0]  # Single file
        if failed > 0:
            print(f"\nSummary: Processing failed ({filename})")
        elif duplicates > 0:
            print(f"\nSummary: Duplicate skipped ({filename})")
        else:
            print(f"\nSummary: Successfully processed ({filename})")
    else:
        print(f"\nSummary: {total_files} files processed ({processed} new, {duplicates} duplicates, {failed} failed)")


if __name__ == '__main__':
    main()